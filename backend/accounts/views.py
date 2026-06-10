from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse
from django.conf import settings
from django.db import connection
from django.middleware.csrf import get_token
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
import re

from django.utils import timezone

from certificates.models import Certificate
from courses.models import Course
from .models import AdminLoginLog, Student
from .serializers import StudentSerializer


def _get_client_ip(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


DEFAULT_PASSWORD = "Tech@123"


def _next_student_id():
    last_student = (
        Student.objects.filter(student_id__startswith="TSC")
        .order_by("-student_id")
        .first()
    )
    if not last_student:
        return "TSC001"

    try:
        next_number = int(last_student.student_id.replace("TSC", "")) + 1
    except ValueError:
        next_number = Student.objects.count() + 1
    return f"TSC{next_number:03d}"


def _ensure_course():
    course, _ = Course.objects.get_or_create(
        course_name="Internship Training",
        defaults={"duration": "3 Months"},
    )
    return course


def _admin_required(request):
    return request.session.get("role") == "admin" or (
        request.user.is_authenticated and request.user.is_staff
    )


@ensure_csrf_cookie
@api_view(["GET"])
def csrf_token(request):
    return Response({"csrfToken": get_token(request)})


@api_view(["POST"])
def login_view(request):
    role = request.data.get("role", "student")

    logout(request)
    request.session.flush()

    if role == "admin":
        username = str(request.data.get("username", "")).strip()
        password = str(request.data.get("password", "")).strip()
        if not username or not password:
            return Response(
                {"error": "Username and password are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user = authenticate(
            request,
            username=username,
            password=password,
        )
        if user and user.is_staff:
            login(request, user)
            request.session["role"] = "admin"

            log = AdminLoginLog.objects.create(
                username=user.username,
                ip_address=_get_client_ip(request),
            )
            request.session["admin_log_id"] = log.id
            request.session.save()

            return Response({
                "message": "Admin login success",
                "role": "admin",
                "username": user.username,
            })
        return Response(
            {"error": "Invalid admin credentials"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    student_id = str(request.data.get("student_id", "")).strip()
    password = str(request.data.get("password", "")).strip()
    if not student_id or not password:
        return Response(
            {"error": "Student ID and password are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        student = Student.objects.get(student_id=student_id)
    except Student.DoesNotExist:
        return Response(
            {"error": "Student not found"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if student.password != password:
        return Response(
            {"error": "Invalid password"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Explicitly create a new session
    if not request.session.session_key:
        request.session.create()
    
    request.session["role"] = "student"
    request.session["student_id"] = student.student_id
    request.session.save()
    
    return Response({
        "message": "Student login success",
        "role": "student",
        "student_id": student.student_id,
        "name": student.name,
    })


@api_view(["POST"])
def logout_view(request):
    log_id = request.session.get("admin_log_id")
    if log_id:
        AdminLoginLog.objects.filter(id=log_id, logout_at__isnull=True).update(
            logout_at=timezone.now()
        )
    logout(request)
    request.session.flush()
    return Response({"message": "Logged out"})


@api_view(["GET"])
def session_view(request):
    role = request.session.get("role")
    return Response({
        "authenticated": bool(role),
        "role": role,
        "student_id": request.session.get("student_id"),
        "is_admin": _admin_required(request),
    })


@api_view(["GET"])
def dashboard_stats(request):
    if not _admin_required(request):
        return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)

    return Response({
        "students": Student.objects.count(),
        "courses": Course.objects.count(),
        "certificates": Certificate.objects.count(),
    })


@api_view(["GET"])
def database_connection(request):
    try:
        db_settings = settings.DATABASES["default"]
        if db_settings.get("ENGINE") == "django_mongodb_backend":
            ping_result = connection.database.command("ping")
            result = ping_result.get("ok")
        else:
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                result = cursor.fetchone()

        return Response({
            "status": "connected",
            "message": "MongoDB connection is healthy.",
            "database": db_settings.get("NAME"),
            "result": result,
        })

    except Exception as e:
        return Response({
            "status": "error",
            "message": str(e),
            "database": settings.DATABASES["default"].get("NAME"),
        }, status=500)


@api_view(["POST"])
def upload_excel(request):
    if not _admin_required(request):
        return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)

    excel_file = request.FILES.get("file")
    if not excel_file:
        return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        from openpyxl import load_workbook
    except ImportError:
        return Response(
            {"error": "OpenPyXL is not installed. Install it to upload Excel files."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    try:
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active
    except Exception:
        return Response(
            {"error": "Invalid Excel file. Upload a valid .xlsx file."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return Response({"error": "Excel file is empty"}, status=status.HTTP_400_BAD_REQUEST)

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_map = {header.lower(): index for index, header in enumerate(headers)}
    if "name" not in header_map or "email" not in header_map:
        return Response(
            {"error": "Excel file must contain Name and Email columns"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    default_course = _ensure_course()
    created = []
    updated = []
    skipped = []
    warnings = []

    # Optional Excel columns:
    # - `course` / `course_name` (preferred): which course the student registered for
    # - `duration`: optional duration used when creating a new course
    course_key = None
    for key in ["course", "course_name", "registered_course", "registered_courses"]:
        if key in header_map:
            course_key = key
            break
    # Fallback: if Excel uses a slightly different header (e.g. "Registered Courses"),
    # pick the first column that contains "course".
    if course_key is None:
        for key in header_map.keys():
            if "course" in key:
                course_key = key
                break
    duration_key = "duration" if "duration" in header_map else None

    def _normalize_course_name(raw):
        """
        Convert Excel cell to a course name.

        IMPORTANT: Keep the full text as-is (trim + collapse whitespace).
        Previously we split by separators which could truncate valid course names.
        """
        if raw is None:
            return ""
        text = str(raw).strip()
        if not text:
            return ""
        # Normalize newlines to spaces and collapse repeated whitespace.
        text = re.sub(r"\s+", " ", text)
        return text

    # Optional student_id column (if Excel provides it).
    student_id_key = None
    for key in header_map.keys():
        k = key.replace(" ", "")
        if k in ["studentid", "student_id"]:
            student_id_key = key
            break
        if "student" in key and "id" in key:
            student_id_key = key
            break

    for row in rows[1:]:
        name_value = row[header_map["name"]] if header_map["name"] < len(row) else ""
        email_value = row[header_map["email"]] if header_map["email"] < len(row) else ""
        name = str(name_value or "").strip()
        email = str(email_value or "").strip().lower()

        if not name or not email:
            skipped.append({"email": email, "reason": "Missing name or email"})
            continue

        raw_course_value = ""
        if course_key and header_map[course_key] < len(row):
            raw_course_value = row[header_map[course_key]]

        selected_course_name = _normalize_course_name(raw_course_value)
        if not selected_course_name:
            selected_course_name = default_course.course_name
            if course_key:
                warnings.append({"email": email, "warning": "Course missing; using default course"})

        raw_duration_value = None
        if duration_key and header_map[duration_key] < len(row):
            raw_duration_value = row[header_map[duration_key]]
        selected_duration = (
            str(raw_duration_value).strip()
            if raw_duration_value is not None and str(raw_duration_value).strip()
            else None
        )

        course_obj, _ = Course.objects.get_or_create(
            course_name=selected_course_name,
            defaults={"duration": selected_duration or default_course.duration},
        )
        if selected_duration and course_obj.duration != selected_duration:
            course_obj.duration = selected_duration
            course_obj.save(update_fields=["duration"])

        existing = Student.objects.filter(email=email).first()
        if existing:
            student_id_changed = False
            existing.name = name or existing.name
            existing.course = course_obj
            if student_id_key and header_map[student_id_key] < len(row):
                desired_student_id = str(row[header_map[student_id_key]] or "").strip()
                if desired_student_id and desired_student_id != existing.student_id:
                    # Avoid breaking uniqueness; only set if it's free.
                    if not Student.objects.filter(student_id=desired_student_id).exists():
                        existing.student_id = desired_student_id
                        student_id_changed = True
                    else:
                        warnings.append(
                            {
                                "email": email,
                                "warning": f"student_id '{desired_student_id}' already exists; keeping existing id",
                            }
                        )
            if student_id_changed:
                existing.save(update_fields=["name", "course", "student_id"])
            else:
                existing.save(update_fields=["name", "course"])
            updated.append(existing.student_id)
            continue

        desired_student_id = None
        if student_id_key and header_map[student_id_key] < len(row):
            desired_student_id = str(row[header_map[student_id_key]] or "").strip()
            if not desired_student_id:
                desired_student_id = None

        if desired_student_id and Student.objects.filter(student_id=desired_student_id).exists():
            warnings.append(
                {
                    "email": email,
                    "warning": f"student_id '{desired_student_id}' already exists; generating new id",
                }
            )
            desired_student_id = None

        final_student_id = desired_student_id or _next_student_id()

        student = Student.objects.create(
            student_id=final_student_id,
            name=name,
            email=email,
            password=DEFAULT_PASSWORD,
            course=course_obj,
        )
        created.append(student.student_id)

    return Response({
        "message": "Students uploaded successfully",
        "created_count": len(created),
        "created_student_ids": created,
        "updated_count": len(updated),
        "updated_student_ids": updated,
        "skipped": skipped,
        "warnings": warnings,
        "default_password": DEFAULT_PASSWORD,
    })


@api_view(["GET", "POST"])
def students(request):
    if not _admin_required(request):
        return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)

    if request.method == "GET":
        student_rows = Student.objects.select_related("course").order_by("student_id")
        return Response(StudentSerializer(student_rows, many=True).data)

    data = request.data.copy()
    data.setdefault("student_id", _next_student_id())
    data.setdefault("password", DEFAULT_PASSWORD)
    if not data.get("course"):
        data["course"] = _ensure_course().id

    serializer = StudentSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET", "PUT", "DELETE"])
def student_detail(request, student_id):
    try:
        student = Student.objects.get(student_id=student_id)
    except Student.DoesNotExist:
        return Response({"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        session_student_id = request.session.get("student_id")
        if not _admin_required(request) and session_student_id != student_id:
            return Response({"error": "Access denied"}, status=status.HTTP_403_FORBIDDEN)
        return Response(StudentSerializer(student).data)

    if not _admin_required(request):
        return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)

    if request.method == "DELETE":
        student.delete()
        return Response({"message": "Student deleted"})

    serializer = StudentSerializer(student, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
def bulk_delete_students(request):
    if not _admin_required(request):
        return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)

    student_ids = request.data.get("student_ids", [])
    if not isinstance(student_ids, list) or not student_ids:
        return Response(
            {"error": "student_ids must be a non-empty list"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    students_qs = Student.objects.filter(student_id__in=student_ids)
    deleted_count = students_qs.count()
    students_qs.delete()
    return Response(
        {
            "message": "Students deleted",
            "deleted_count": deleted_count,
        }
    )


@api_view(["GET"])
def student_profile(request, student_id):
    try:
        student = Student.objects.get(student_id=student_id)
    except Student.DoesNotExist:
        return Response({"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND)

    # Check access control: admin can view any, students can only view their own
    if not _admin_required(request):
        session_student_id = request.session.get("student_id")
        if not session_student_id or session_student_id != student_id:
            return Response({"error": "Access denied"}, status=status.HTTP_403_FORBIDDEN)

    return Response(StudentSerializer(student).data)


@api_view(["GET"])
def admin_login_logs(request):
    if not _admin_required(request):
        return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)

    logs = AdminLoginLog.objects.all()[:100]
    data = [
        {
            "id": log.id,
            "username": log.username,
            "login_at": log.login_at,
            "logout_at": log.logout_at,
            "ip_address": log.ip_address,
        }
        for log in logs
    ]
    return Response(data)
